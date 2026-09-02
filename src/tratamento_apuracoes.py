import pandas as pd
from python_calamine import CalamineWorkbook
import time
import datetime
import glob
from openpyxl import load_workbook
import unicodedata
import logging
import os

# Variável global dos restaurantes que tem almoço e jantar
ABREM_TODO_DIA_ALMOCO_E_JANTAR = {
    "ACIARIA SUL", "COQUERIA", "MANUTENÇÃO CENTRAL", "MINI CONTÍNUO", 
    "MINI CONVERTEDOR", "MINI LTQ", "SUNCOKE", "CENTRAL"
}

# ============================================================================== #
#                       Funções auxiliares e de análise                          #
# ============================================================================== #

def encontrar_arquivo_apuracao():
    # Busca o arquivo de apuração na pasta raiz.
    arquivos_excel = glob.glob('*.xlsx')
    for nome_arquivo in arquivos_excel:
        if "apuracao_geral_arcelormittal" in nome_arquivo.lower():
            return nome_arquivo
    return None 

def gerar_intervalo_de_datas(data_inicial, data_final):
    # Gera uma lista de strings de data entre as datas inicial e final.
    data_inicial = datetime.datetime.strptime(data_inicial, '%d/%m/%Y')
    data_final = datetime.datetime.strptime(data_final, '%d/%m/%Y')
    
    intervalo = []
    while data_inicial <= data_final:
        intervalo.append(data_inicial.strftime('%d/%m/%Y')) 
        data_inicial += datetime.timedelta(days=1)
    
    return intervalo

def normalizar_texto(texto):
    texto = texto.upper()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto.strip()

def obter_data():
    while True:
        try:
            data_inicial = input("\n>>> Digite a data inicial (ex: 01/01/2001): ")
            data_inicial_formatada = datetime.datetime.strptime(data_inicial, '%d/%m/%Y')
            data_inicial_intervalo = data_inicial_formatada.strftime('%d/%m/%Y')

            deseja_intervalo = str(input("\n>>> Deseja filtrar por intervalo? (S/N): ")).lower()

            if deseja_intervalo == "s":
                data_final = input("\n>>> Insira a data final (ex: 01/01/2001): ")
                data_final_formatada = datetime.datetime.strptime(data_final, '%d/%m/%Y')
                data_final_intervalo = data_final_formatada.strftime('%d/%m/%Y')

                print(f"\nFiltrando pelo intervalo de {data_inicial_intervalo} a {data_final_intervalo}!\n")
                return gerar_intervalo_de_datas(data_inicial_intervalo, data_final_intervalo)
            else:
                print(f"\nFiltrando pela data {data_inicial_intervalo}!\n")
                return [data_inicial_intervalo]
        except ValueError:
            print("Formato ou intervalo inválidos. Utilize o formato 'dia/mes/ano'.\n")

def formatar_coluna_data(caminho_arquivo, nome_coluna='data', nome_aba=None):
    # Formata a coluna de data no arquivo Excel final usando openpyxl
    try:
        wb = load_workbook(caminho_arquivo)
        if nome_aba and nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
        else:
            ws = wb.active

        col_index = None
        for cell in ws[1]:
            if cell.value == nome_coluna:
                col_index = cell.column
                break
        
        if not col_index:
            print(f"Aviso: coluna '{nome_coluna}' não encontrada na aba '{ws.title}'.")
            return

        for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
            cell = row[0]
            try:
                data_obj = pd.to_datetime(cell.value, format='%d/%m/%Y', errors='coerce')
                if pd.notna(data_obj):
                    cell.value = data_obj
                    cell.number_format = 'DD/MM/YYYY'
            except (ValueError, TypeError):
                continue

        wb.save(caminho_arquivo)
        print(f"Formatação da coluna '{nome_coluna}' aplicada na aba '{ws.title}'.")

    except Exception as e:
        print(f"ERRO: Ocorreu um erro ao formatar a planilha: {e}")

def definir_categoria_preparacao(preparacao):
    # Define a categoria (PROTEINA, SALADA, ARROZ, etc.) com base no nome do produto.
    preparacao = str(preparacao).upper().strip()
    categorias = {
        "GUARNICAO": ["ESPAGUETADA", "CREME", "POLENTA", "FAROFA", "QUIBEBE", "CANJIQUINHA", "PENNE", "PURE", "ESPAGUETE", "MACARRAO", 
                        "VIRADO", "CUSCUZ", "PIRAO", "NHOQUE", "PALHA"],
        "SALADA": ["SAL.", "ALFACE", "BETERRABA", "BERINGELA", "BERINJELA", "LENTILHA", "PEPINO", "TOMATE", "VAGEM", "CENOURA", "ERVILHA", 
                     "CHUCHU", "ABOBORA", "BATATA", "LEGUMES", "LEGUME", "COUVE", "SOJA", "REPOLHO", "TRIGO", "JILO", "GRAO", "BROCOLIS",
                     "ABOBRINHA", "JARDINEIRA"],
        "PROTEINA": ["COZIDO", "FRANGO", "BIFE", "FILEZINHO", "KIBINHO", "LINGUICA", "OVOS", "MERLUZA", "STROGONOFF", "CARRE", "ATUM", 
                       "CARNE", "OMELETE", "SALSICHA", "SALSICHAO", "ISCAS", "TILAPIA", "BISTECA", "HAMBURGUER", "EMPADAO", "PERNIL", "PICADINHO", 
                       "QUIBE", "FRICASSE", "BOBO", "CUBOS", "FILE", "ALMONDEGA", "ALMONDEGAS", "LOMBO", "DOBRADINHA", "GOULASH", "QUICHE", "KIBE",
                       "MOQUECA", "MOUSSAKA", "COSTELA", "FEIJOADA", "CORDON", "LASANHA", "MOELA", "OVO", "SOBRECOXA", "PATINHO", "FIGADO", 
                       "PIZZA", "COSTELINHA"],
        "SOBREMESA": ["DOCE", "MACA", "MELANCIA", "MELAO", "CHAMOUR", "LARANJA", "DELICIA", "MANJAR", "PUDIM", "TORTA", "CURAU", "GOIABADA",
                       "CHOCOLATE", "FLAN", "PAVE", "CACAROLA", "GELATINA", "PERA", "BANANADA", "BANANA", "MAMAO", "COCADA", "PE", "PICOLE", 
                       "ABACAXI", "TANGERINA", "TIRAMISSU", "BARRA"]
    }
    categorias_fixas = {"ARROZ": "ARROZ", "FEIJAO": "FEIJAO", "SUCO": "SUCO", "MOLHO": "MOLHO", "Z": preparacao}

    lista_preparacao = preparacao.split()
    if not lista_preparacao: return None 
        
    prim_nome_preparacao = lista_preparacao[0]
    
    if prim_nome_preparacao in categorias_fixas:
        return categorias_fixas[prim_nome_preparacao]

    if len(lista_preparacao) >= 2 and lista_preparacao[1] == "PALHA":
        return "GUARNICAO"

    for categoria, palavras in categorias.items():
        if prim_nome_preparacao in palavras:
            return categoria

    return None

def criar_logger(data_log):
    os.makedirs("resumos_apuracao", exist_ok=True)
    nome_arquivo = f"resumos_apuracao/apuracao_{data_log}.log"

    logger = logging.getLogger(f"resumo_{data_log}")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter(
        "%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%d/%m/%Y %H:%M:%S"
    )

    file_handler = logging.FileHandler(nome_arquivo, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.propagate = False

    return logger

def gerar_resumo_pesagens(df, logger, etapas_mestre, restaurantes_mestre):
    logger.info("\n")
    logger.info("=" * 90)
    logger.info("RESUMO DE PESAGENS POR RESTAURANTE E ETAPA")
    logger.info("=" * 90 + "\n")

    etapas_df = pd.DataFrame({'etapa': etapas_mestre})

    resumo = (
        df.groupby(['restaurante', 'etapa'])['pesagem']
        .sum()
        .reset_index()
    )
    resumo['etapa'] = resumo['etapa'].str.upper()

    for restaurante in sorted(restaurantes_mestre):
        logger.info("-" * 90)
        logger.info(f"RESTAURANTE: {restaurante}")
        logger.info("-" * 90)

        resumo_rest = resumo[resumo['restaurante'] == restaurante]

        resumo_completo = (
            etapas_df
            .merge(resumo_rest[['etapa', 'pesagem']], on='etapa', how='left')
            .fillna({'pesagem': 0.0})
        )

        for _, row in resumo_completo.iterrows():
            logger.info(f"{row['etapa']:<40} : {row['pesagem']:.2f} kg")

        logger.info("")

def definir_etapa(etapa, restaurante, balanca, produto):
    e = normalizar_texto(etapa)

    RESTAURANTES = [
        "ACIARIA SUL", "ACIARIA NORTE", "ACABAMENTO", "ALTO FORNO",
        "CENTRO", "COQUERIA", "SUNCOKE", "MINI CONTINUO", "MINI LTQ",
        "MINI CONVERTEDOR", "MANUTENCAO", "SINTERIZACAO", "TRANSPORTE",
    ]

    if "RESTO INGESTA" in produto:
        return "RESTO INGESTA"
    if "SOBRA LIMPA" in e or "SOB LIMPA" in e:
        return "SOBRA LIMPA"
    if "CADENCIAMENTO" in e:
        if "CENTRAL" in e:
            return "CADENCIAMENTO CENTRAL"
        return "CADENCIAMENTO"
    if balanca in ["CENTRAL", "CENTRAL SALADA", "CENTRAL CONFEITARIA"]:
        if any(r in e for r in RESTAURANTES):
            return "PRODUCAO INICIAL TRANSPORTADA"
        else:
            return "PRODUCAO INICIAL CENTRAL"
    if "PRODUCAO" in e:
        return "PRODUCAO INICIAL"
    if "PERDA" in e:
        if "ARMAZENAMENTO" in e:
            return "PERDA ARMAZENAMENTO"
        return "PERDA POR PREPARACAO"
    if "ENTRADA REQUISICAO" in e:
        return "ENTRADA REQUISICAO"
    if "REQUISICAO" in e:
        return "REQUISICAO"
    if "EXTRA REQUISICAO" in e:
        return "EXTRA REQUISICAO"
    if "RECEBIMENTO" in e:
        return "RECEBIMENTO"
    if "ENTRADA DE PRODUTO" in e:
        return "ENTRADA DE PRODUTO"
    if "DEVOLUCAO" in e:
        return "DEVOLUCAO QUALIDADE"
    if "PROTEINA PROCESSADA" in e:
        return "PROTEINA PROCESSADA"
    if "PROTEINA CONGELADA" in e:
        return "PROTEINA CONGELADA"
    if "APARAS" in e:
        return "APARAS"
    if "SOCORRO" in e:
        return "SOCORRO"
    if "ANTECIP" in e:
        return "ANTECIPACAO"
    if "REGENER" in e:
        return "REGENERACAO"

    return etapa

def definir_turno_da_pesagem(restaurante, horario, etapa):
    etapa_upper = etapa.upper().strip() 

    if "ALM" in etapa_upper:
        return "ALMOCO"
    if "JAN" in etapa_upper:
        return "JANTAR"
    
    abre_todo_dia_almoco_e_jantar = restaurante in ABREM_TODO_DIA_ALMOCO_E_JANTAR
    eh_prod_inicial = "PRODUCAO INICIAL" in etapa_upper
    eh_prod_transportada = "TRANSPORTADA" in etapa_upper
    eh_resto_ingesta = "RESTO INGESTA" in etapa_upper

    if eh_resto_ingesta:
        if horario > "14:00:00" or horario < "19:00:00":
            return "ALMOCO"
        return "JANTAR"
    
    if not abre_todo_dia_almoco_e_jantar:
        return "ALMOCO"
    
    if eh_prod_inicial and eh_prod_transportada:
        if horario > "16:00:00" or horario < "06:00:00":
            return "ALMOCO"
        return "JANTAR"

    if eh_prod_inicial:
        if horario > "03:00:00" and horario < "13:30:00":
            return "ALMOCO"
        return "JANTAR"
    
    if horario > "06:00:00" and horario < "17:00:00":
        return "ALMOCO"
    return "JANTAR"

def extrair_df_aba(wb, nome_aba):
    """Extrai uma aba via Calamine e retorna como DataFrame."""
    try:
        planilha = wb.get_sheet_by_name(nome_aba)
        dados = planilha.to_python()
        if not dados or len(dados) < 2:
            return pd.DataFrame()
        return pd.DataFrame(dados[1:], columns=dados[0])
    except Exception:
        return pd.DataFrame()

# ============================================================================== #
#                               Função principal                                 #
# ============================================================================== #

arquivo_entrada = encontrar_arquivo_apuracao()

if arquivo_entrada:
    arquivo_saida = f"apuracao_consolidada_{arquivo_entrada[37:].replace('.xlsx','')}.xlsx"
else:
    arquivo_saida = "apuracao_consolidada_ERRO.xlsx"

def tratar_planilha_apuracao():
    if not arquivo_entrada:
        print("ERRO: Nenhum arquivo .xlsx com 'apuracao_geral_arcelormittal' foi encontrado na pasta raiz.")
        input()
        return

    dfs = []
    
    try:
        wb = CalamineWorkbook.from_path(arquivo_entrada)
        print("=========================================================================================")
        print("#\tTRATAMENTO APURAÇÃO DE PESAGENS BALANÇAS IOS - SAPORE ARCELORMITTAL TUBARÃO\t#")
        print(f"#\t\tEm caso de dúvidas ou sugestões, romulo.santana@sapore.com.br\t\t#")     
        print("=========================================================================================\n")
   
        while True:
            opcao_usuario = input(">>> Deseja filtrar por data? (S/N): ").lower()
            if opcao_usuario == 's':
                data_para_filtro = obter_data()
                break
            elif opcao_usuario == 'n':
                print("\n\nProcessando todas as datas. Aguarde...\n\n")
                data_para_filtro = None
                break
            else:
                print("Opção inválida. Tente novamente.")
                continue

        nomes_abas_disponiveis = wb.sheet_names

        # Carregamento e inserção de colunas dependentes da aba de pesagens
        for aba_nome in nomes_abas_disponiveis:
            inicio = time.time()
            if "3352 - " not in aba_nome.upper(): continue

            print(f"Tratando aba {aba_nome}.")
            nome_planilha = wb.get_sheet_by_name(aba_nome)
            dados_wb = nome_planilha.to_python()

            if not dados_wb or len(dados_wb) < 2:
                print(f"Aba '{aba_nome}' está vazia ou contém apenas cabeçalho.")
                continue

            df = pd.DataFrame(dados_wb[1:], columns=dados_wb[0])
            colunas_a_manter = ["data", "horario", "etapa", "produto", "panela", "pesagem", "servico"]
            df = df[[col for col in colunas_a_manter if col in df.columns]]

            # Inserção de restaurante e balança
            nome_balanca = aba_nome.replace("3352 - ", "")
            nome_restaurante = aba_nome.replace("3352 - ", "").replace(" RECEB", "").replace(" HIB", "")
            if nome_restaurante in ["CENTRAL SOBRA LIMPA", "CENTRAL ACOUGUE", "CENTRAL CONFEITARIA", "CENTRAL SALADA", "CENTRAL ESTOQUE"]:
                nome_restaurante = "CENTRAL"
            
            df.insert(loc=1, column="restaurante", value=nome_restaurante)
            df.insert(loc=2, column="balanca", value=nome_balanca)

            # Formatação da coluna de data na planilha
            if 'data' in df.columns and not df['data'].empty:
                try:
                    df['data'] = pd.to_datetime(df['data'], errors='coerce')
                    df['data'] = df['data'].dt.strftime('%d/%m/%Y')
                    df.dropna(subset=['data'], inplace=True)
                except Exception as e:
                    print(f"Não foi possível formatar a coluna 'data' na planilha '{aba_nome}'. Erro: {e}")
            
            if opcao_usuario == 's':
                df = df[df['data'].isin(data_para_filtro)]

            if not df.empty:
                dfs.append(df)
            else:
                print(f"Aba '{aba_nome}' resultou em um DataFrame vazio após o processamento/filtragem")

        # ====================================================================== #
        #       Tratamento das Abas: NOTA ITEM e NOTA CONFERENCIA                #
        # ====================================================================== #
        df_nota_item = pd.DataFrame()
        df_nota_conferencia = pd.DataFrame()

        # 1. NOTA ITEM
        if "NOTA ITEM" in nomes_abas_disponiveis:
            print("\nTratando aba NOTA ITEM...")
            df_ni = extrair_df_aba(wb, "NOTA ITEM")
            if not df_ni.empty:
                colunas_nota_item = [
                    "dt_emissao", "num_nota", "chave_nota", "categoria_1", "categoria_2",
                    "produto_estoque", "conferido", "qtde_nota", "qtde_contagem"
                ]
                df_nota_item = df_ni[[col for col in colunas_nota_item if col in df_ni.columns]].copy()

                if 'dt_emissao' in df_nota_item.columns:
                    try:
                        df_nota_item['dt_emissao'] = pd.to_datetime(df_nota_item['dt_emissao'], errors='coerce').dt.strftime('%d/%m/%Y')
                    except Exception as e:
                        print(f"Erro ao processar dt_emissao em NOTA ITEM: {e}")

                if opcao_usuario == 's' and 'dt_emissao' in df_nota_item.columns:
                    df_nota_item = df_nota_item[df_nota_item['dt_emissao'].isin(data_para_filtro)]
            else:
                print("Aba 'NOTA ITEM' está vazia.")

        # 2. NOTA CONFERENCIA
        if "NOTA CONFERENCIA" in nomes_abas_disponiveis:
            print("Tratando aba NOTA CONFERENCIA...")
            df_nc = extrair_df_aba(wb, "NOTA CONFERENCIA")
            if not df_nc.empty:
                colunas_nota_conf = [
                    "dt_emissao", "num_nota", "chave_nota", "status_nota",
                    "conferencia_final", "qtde_itens_nota"
                ]
                df_nota_conferencia = df_nc[[col for col in colunas_nota_conf if col in df_nc.columns]].copy()

                if 'dt_emissao' in df_nota_conferencia.columns:
                    try:
                        df_nota_conferencia['dt_emissao'] = pd.to_datetime(df_nota_conferencia['dt_emissao'], errors='coerce').dt.strftime('%d/%m/%Y')
                    except Exception as e:
                        print(f"Erro ao processar dt_emissao em NOTA CONFERENCIA: {e}")

                if opcao_usuario == 's' and 'dt_emissao' in df_nota_conferencia.columns:
                    df_nota_conferencia = df_nota_conferencia[df_nota_conferencia['dt_emissao'].isin(data_para_filtro)]
            else:
                print("Aba 'NOTA CONFERENCIA' está vazia.")

        # INSERÇÃO DE COLUNAS NO DATAFRAME COM BASE EM FUNÇÕES
        dfs_validos = [df for df in dfs if not df.empty]
        if dfs_validos:
            df_final = pd.concat(dfs_validos, ignore_index=True)

            if opcao_usuario == 's':
                datas_unicas = data_para_filtro
            else:
                datas_unicas = sorted(df_final['data'].unique())

            if 'horario' in df_final.columns:
                 df_final['horario'] = df_final['horario'].astype(str)
            else:
                 raise Exception("Dados sem coluna 'horario'.")
            
            # Coluna de turno
            df_final.insert(loc=3, column="turno", value=df_final.apply(
                lambda row: definir_turno_da_pesagem(restaurante=row['restaurante'], horario=row['horario'], etapa=row['etapa']), axis=1
            ))
            
            # Coluna de categoria
            if 'produto' in df_final.columns:
                df_final.insert(loc=6, column="categoria", value=df_final['produto'].apply(definir_categoria_preparacao))
            else:
                 df_final.insert(loc=6, column="categoria", value=None)

            # Tratamento da coluna de etapa
            df_final["etapa"] = df_final.apply(
                lambda row: definir_etapa(
                    etapa=row["etapa"],
                    restaurante=row["restaurante"],
                    balanca=row["balanca"],
                    produto=row["produto"]
                ),
                axis=1
            )
            
            colunas_finais = [
                'data', 'restaurante', 'turno', 'balanca', 'horario', 
                'categoria', 'etapa', 'produto', 'panela', 'pesagem', 
                'servico',
            ]

            df_final = df_final[[c for c in colunas_finais if c in df_final.columns]]

            # Salvamento das abas consolidadas
            print(f"\nGravando abas no arquivo {arquivo_saida}...")
            with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='pesagens', index=False)
                if not df_nota_item.empty:
                    df_nota_item.to_excel(writer, sheet_name='nota_item', index=False)
                if not df_nota_conferencia.empty:
                    df_nota_conferencia.to_excel(writer, sheet_name='nota_conferencia', index=False)

            # Formatação de datas em cada aba
            print("\nAplicando formatações de data com openpyxl...")
            formatar_coluna_data(arquivo_saida, nome_coluna='data', nome_aba='pesagens')
            if not df_nota_item.empty:
                formatar_coluna_data(arquivo_saida, nome_coluna='dt_emissao', nome_aba='nota_item')
            if not df_nota_conferencia.empty:
                formatar_coluna_data(arquivo_saida, nome_coluna='dt_emissao', nome_aba='nota_conferencia')
            if not df_nota_conferencia.empty:
                formatar_coluna_data(arquivo_saida, nome_coluna='conferencia_final', nome_aba='nota_conferencia')

            fim = time.time()
            print(f"\nTotal de linhas de pesagens processadas: {len(df_final)}")
            print(f"Total de linhas em nota_item: {len(df_nota_item)}")
            print(f"Total de linhas em nota_conferencia: {len(df_nota_conferencia)}")
            print(f"Tempo de execução: {(fim - inicio):.2f} segundos")

            etapas_mestre_global = (
                df_final['etapa']
                .dropna()
                .str.upper()
                .unique()
                .tolist()
            )

            restaurantes_mestre_global = (
                df_final['restaurante']
                .dropna()
                .unique()
                .tolist()
            )

            for data_log in datas_unicas:
                logger = criar_logger(data_log.replace("/", "-"))

                df_dia = df_final[df_final['data'] == data_log]

                logger.info(f"Arquivo de consolidação: {arquivo_saida}")
                logger.info(f"Data processada: {data_log}")
                logger.info(f"Total de linhas processadas: {len(df_dia)}")
                logger.info(f"Total de planilhas processadas: {len(dfs_validos)}")
                logger.info(f"Tempo de execução: {(fim - inicio):.2f} segundos")

                gerar_resumo_pesagens(df_dia, logger, etapas_mestre_global, restaurantes_mestre_global)
            print("Resumo das pesagens por dia criados, confira na pasta 'resumos_apuracao'")

        else:
            print("Nenhuma aba válida para consolidação.")
    except PermissionError:
        print("\nErro ao sobrescrever o arquivo. Feche a planilha e tente novamente.")
    except Exception as e:
        print(f"\nErro durante a execução: {e}")
        
    input("\nPressione ENTER para sair.")

if __name__ == "__main__":
    tratar_planilha_apuracao()